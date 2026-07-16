from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path
from datetime import date

from database import get_attendance_list, get_days

BASE_DIR = Path(__file__).resolve().parent
REPORTS_DIR = BASE_DIR / "attendance_reports"

HEADER = ["Employee ID", "Role", "Timestamp", "Latitude", "Longitude", "Address"]


def export_month(year=None, month=None):
    """One Excel file for the month with a Summary tab + one tab per day.

    File: attendance_reports/attendance_YYYY-MM.xlsx
    """
    today = date.today()
    year = year or today.year
    month = month or today.month
    ym = f"{year:04d}-{month:02d}"

    REPORTS_DIR.mkdir(exist_ok=True)
    EXCEL_PATH = REPORTS_DIR / f"attendance_{ym}.xlsx"

    days = get_days(ym)  # dates in this month that have attendance

    wb = Workbook()

    # ---- Summary / title sheet ----
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = f"Attendance Summary - {ym}"
    summary["A1"].font = Font(bold=True, size=14)
    summary.append([])
    summary.append(["Date", "Employees Present"])
    summary["A3"].font = Font(bold=True)
    summary["B3"].font = Font(bold=True)

    # ---- One sheet per day ----
    for d in days:
        rows = get_attendance_list("employee", d)
        summary.append([d, len(rows)])

        ws = wb.create_sheet(title=d)  # tab name = the date, e.g. 2026-07-16
        ws.append(HEADER)
        for c in ws[1]:
            c.font = Font(bold=True)
        for row in rows:
            ws.append([row[0], "Employee", row[1], row[2], row[3], row[4]])

    wb.save(EXCEL_PATH)
    print(f"[export] Monthly workbook updated: {EXCEL_PATH}")
    return EXCEL_PATH


def export_today_data():
    today = date.today().isoformat()

    REPORTS_DIR.mkdir(exist_ok=True)
    EXCEL_PATH = REPORTS_DIR / f"attendance_{today}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance {today}"

    ws.append(["Employee ID", "Role", "Timestamp", "Latitude", "Longitude", "Address"])

    for row in get_attendance_list("employee", today):
        ws.append([row[0], "Employee", row[1], row[2], row[3], row[4]])

    wb.save(EXCEL_PATH)

    print(f"[export] Excel updated: {EXCEL_PATH}")
    return EXCEL_PATH


def export_all_data():
    """Export the FULL attendance history (all days) into one Excel file."""
    REPORTS_DIR.mkdir(exist_ok=True)
    EXCEL_PATH = REPORTS_DIR / "attendance_full_report.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "All Attendance"

    ws.append(["Employee ID", "Role", "Timestamp", "Latitude", "Longitude", "Address"])

    for row in get_attendance_list("employee", None):
        ws.append([row[0], "Employee", row[1], row[2], row[3], row[4]])

    wb.save(EXCEL_PATH)

    print(f"[export] Full report written: {EXCEL_PATH}")
    return EXCEL_PATH
